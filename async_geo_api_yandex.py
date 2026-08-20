import os
import asyncio
import aiohttp
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from math import radians, sin, cos, sqrt, atan2
import time

# API_KEY = os.getenv('API_KEY')
API_KEY = "640e607e-14bb-42c8-8ffd-7aa33f5e587a"
TORGER_ADDRESS = "Москва, Рязанский проспект, 99"
FILENAME = "fake_data.xlsx"

async def geocode_address(session, address, api_key):
    try:
        url = "https://geocode-maps.yandex.ru/1.x/"
        params = {
            "apikey": api_key,
            "geocode": address,
            "format": "json"
        }
        
        async with session.get(url, params=params, timeout=10) as response:
            data = await response.json()
            
            feature_members = data['response']['GeoObjectCollection']['featureMember']
            if not feature_members:
                return None
            
            point = feature_members[0]['GeoObject']['Point']['pos']
            lon, lat = point.split()
            return float(lat), float(lon)
    
    except (aiohttp.ClientError, KeyError, IndexError, ValueError, asyncio.TimeoutError) as e:
        print(f"Ошибка при геокодировании адреса '{address}': {e}")
        return None

def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    return R * c

async def process_student(session, row, address, coords_torger, api_key):
    if not address or str(address).strip() == "":
        return {
            'row': row,
            'distance': "Адрес не указан",
            'status': 'empty'
        }
    
    coords = await geocode_address(session, address, api_key)
    
    if not coords:
        return {
            'row': row,
            'distance': "Адрес не найден",
            'status': 'not_found'
        }
    
    dist = haversine_distance(
        coords_torger[0], coords_torger[1],
        coords[0], coords[1]
    )
    
    return {
        'row': row,
        'distance': round(dist, 2),
        'status': 'success'
    }

async def distance_async():
    async with aiohttp.ClientSession() as session:
        coords_torger = await geocode_address(session, TORGER_ADDRESS, API_KEY)
        if not coords_torger:
            print("Не удалось получить координаты офиса. Скрипт остановлен.")
            return
    
    wb = load_workbook(FILENAME)
    ws = wb.active
    
    address_column = None
    distance_column = None
    
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value and "адрес" in str(cell_value).lower() and "регистр" in str(cell_value).lower():
            address_column = col
        elif cell_value and "расстоян" in str(cell_value).lower():
            distance_column = col
    
    if not distance_column:
        distance_column = ws.max_column + 1
        ws.cell(row=1, column=distance_column, value="Расстояние(км)")
        ws.cell(row=1, column=distance_column).font = Font(bold=True)
        ws.cell(row=1, column=distance_column).alignment = Alignment(horizontal='center')
    
    students = []
    for row in range(2, ws.max_row + 1):
        address = ws.cell(row=row, column=address_column).value
        students.append({
            'row': row,
            'address': address
        })
    
    semaphore = asyncio.Semaphore(10)
    
    async def process_with_semaphore(session, student):
        async with semaphore:
            return await process_student(
                session, 
                student['row'], 
                student['address'], 
                coords_torger, 
                API_KEY
            )
    
    start_time = time.time()
    
    async with aiohttp.ClientSession() as session:
        tasks = [process_with_semaphore(session, student) for student in students]
        results = await asyncio.gather(*tasks)
    
    for result in results:
        row = result['row']
        ws.cell(row=row, column=distance_column, value=result['distance'])
        ws.cell(row=row, column=distance_column).alignment = Alignment(horizontal='center')
    
    wb.save(FILENAME)
    
    total = len(results)
    success = sum(1 for r in results if r['status'] == 'success')
    empty = sum(1 for r in results if r['status'] == 'empty')
    not_found = sum(1 for r in results if r['status'] == 'not_found')
    
    elapsed_time = time.time() - start_time
    
    print(f"Обработано строк: {total}")
    print(f"Успешно обработано: {success}")
    print(f"Пустых адресов: {empty}")
    print(f"Адресов не найдено: {not_found}")
    print(f"Время выполнения: {elapsed_time:.2f} секунд")
    print(f"Результат сохранен в {FILENAME}")

def distance():
    asyncio.run(distance_async())

# if __name__ == "__main__":
distance()