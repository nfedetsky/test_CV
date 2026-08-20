import os
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from math import radians, sin, cos, sqrt, atan2
import time

# API_KEY = os.getenv('API_KEY')
API_KEY = "640e607e-14bb-42c8-8ffd-7aa33f5e587a"
TORGER_ADDRESS = "Москва, Рязанский проспект, 99"
FILENAME = "fake_data.xlsx"

def geocode_address(address, api_key):
    try:
        url = "https://geocode-maps.yandex.ru/1.x/"
        params = {
            "apikey": api_key,
            "geocode": address,
            "format": "json"
        }
        
        response = requests.get(url, params=params, timeout=10)
        data = response.json()
        
        feature_members = data['response']['GeoObjectCollection']['featureMember']
        if not feature_members:
            return None
        
        point = feature_members[0]['GeoObject']['Point']['pos']
        lon, lat = point.split()
        return float(lat), float(lon)
    
    except (requests.RequestException, KeyError, IndexError, ValueError) as e:
        print(f"Ошибка при геокодировании адреса '{address}': {e}")
        return None

def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    return R * c

def distance():
    start_time = time.time()

    coords_torger = geocode_address(TORGER_ADDRESS, API_KEY)
    if not coords_torger:
        print("Не удалось получить координаты университета. Скрипт остановлен.")
        return
    
    wb = load_workbook(FILENAME)
    ws = wb.active
    
    address_column = None
    distance_column = None
    
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value and "адрес" in str(cell_value).lower() and "регистр" in str(cell_value).lower():
            address_column = col
        elif cell_value and "расстоян" in str(cell_value).lower():
            distance_column = col
    
    if not distance_column:
        distance_column = ws.max_column + 1
        ws.cell(row=1, column=distance_column, value="Расстояние(км)")
        ws.cell(row=1, column=distance_column).font = Font(bold=True)
        ws.cell(row=1, column=distance_column).alignment = Alignment(horizontal='center')
    
    error_count = 0
    empty_address_count = 0
    
    for row in range(2, ws.max_row + 1):
        address = ws.cell(row=row, column=address_column).value
        
        if not address or str(address).strip() == "":
            ws.cell(row=row, column=distance_column, value="Адрес не указан")
            ws.cell(row=row, column=distance_column).alignment = Alignment(horizontal='center')
            empty_address_count += 1
            continue
        
        coords = geocode_address(address, API_KEY)
        
        if not coords:
            ws.cell(row=row, column=distance_column, value="Адрес не найден")
            ws.cell(row=row, column=distance_column).alignment = Alignment(horizontal='center')
            error_count += 1
            continue
        
        dist = haversine_distance(
            coords_torger[0], coords_torger[1],
            coords[0], coords[1]
        )
        
        ws.cell(row=row, column=distance_column, value=round(dist, 2))
        ws.cell(row=row, column=distance_column).alignment = Alignment(horizontal='center')
    
    wb.save(FILENAME)
    
    total_rows = ws.max_row - 1
    elapsed_time = time.time() - start_time

    print(f"Обработано строк: {total_rows}")
    print(f"Пустых адресов: {empty_address_count}")
    print(f"Адресов не найдено: {error_count}")
    print(f"Успешно обработано: {total_rows - empty_address_count - error_count}")
    print(f"Время выполнения: {elapsed_time:.2f} секунд")
    print(f"Результат сохранен в {FILENAME}")

# if __name__ == "__main__":
distance()